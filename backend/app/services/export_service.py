"""Сервис экспорта — генерит Excel-файлы с результатами скрапинга.

Используем openpyxl — он ничего не весит и умеет все что нужно.
"""

from __future__ import annotations

from io import BytesIO
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.logging import get_logger
from app.models.task_run import TaskRun

if TYPE_CHECKING:
    from app.models.scraping_task import ScrapingTask

log = get_logger(__name__)

# стили для заголовков — чтоб было красиво
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")


def _style_header(ws, row: int = 1) -> None:
    """Накатывает стили на заголовочную строку — синий фон, белый текст."""
    for cell in ws[row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT


def _auto_width(ws) -> None:
    """Автоподбор ширины столбцов — чтоб текст не обрезался."""
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except (TypeError, AttributeError):
                pass
        adjusted_width = min(max_length + 4, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


async def export_task_excel(
    task: ScrapingTask,
    runs: list[TaskRun] | None,
    db: AsyncSession,
) -> BytesIO:
    """Экспорт данных одной задачи в Excel — один лист с историей запусков."""
    wb = Workbook()
    ws = wb.active
    ws.title = task.name[:31]  # ограничение Excel на длину имени листа

    # достаем запуски с результатами если не переданы
    if runs is None:
        runs_result = await db.execute(
            select(TaskRun)
            .options(selectinload(TaskRun.results))
            .where(TaskRun.task_id == task.id, TaskRun.status == "success")
            .order_by(TaskRun.started_at.desc())
        )
        runs = list(runs_result.scalars().unique().all())

    # собираем список имен полей из первого запуска (или из задачи)
    field_names: list[str] = []
    if runs and runs[0].results:
        seen = set()
        for r in runs[0].results:
            if r.field_name not in seen:
                field_names.append(r.field_name)
                seen.add(r.field_name)
    elif task.fields:
        field_names = [f.name for f in task.fields]

    # заголовки
    headers = ["Дата запуска", "Статус", "Длительность (мс)"] + field_names
    ws.append(headers)
    _style_header(ws)

    # данные по каждому запуску
    for run in runs:
        row_data: list[str] = [
            run.started_at.strftime("%Y-%m-%d %H:%M:%S") if run.started_at else "",
            run.status,
            str(run.duration_ms or ""),
        ]

        # собираем результаты в словарь для быстрого доступа
        results_map: dict[str, str] = {}
        if run.results:
            for r in run.results:
                if r.field_index == 0:
                    results_map[r.field_name] = r.field_value or ""
                else:
                    # для списковых полей конкатенируем через запятую
                    existing = results_map.get(r.field_name, "")
                    results_map[r.field_name] = f"{existing}, {r.field_value}" if existing else (r.field_value or "")

        for fname in field_names:
            row_data.append(results_map.get(fname, ""))

        ws.append(row_data)

    _auto_width(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


async def export_summary_excel(
    tasks: list[ScrapingTask],
    db: AsyncSession,
) -> BytesIO:
    """Экспорт сводки по всем задачам — общая статистика в одном файле."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Сводка"

    # заголовки сводки
    headers = [
        "Название задачи",
        "URL",
        "Статус",
        "Активна",
        "Всего запусков",
        "Успешных",
        "Последний запуск",
        "Расписание",
    ]
    ws.append(headers)
    _style_header(ws)

    for task in tasks:
        ws.append(
            [
                task.name,
                task.url,
                task.last_status or "—",
                "Да" if task.is_active else "Нет",
                task.total_runs,
                task.success_runs,
                task.last_run_at.strftime("%Y-%m-%d %H:%M:%S") if task.last_run_at else "—",
                task.schedule_cron,
            ]
        )

    _auto_width(ws)

    # дополнительные листы для каждой задачи с последними результатами
    for task in tasks:
        # берем последний успешный запуск
        run_result = await db.execute(
            select(TaskRun)
            .options(selectinload(TaskRun.results))
            .where(TaskRun.task_id == task.id, TaskRun.status == "success")
            .order_by(TaskRun.started_at.desc())
            .limit(1)
        )
        last_run = run_result.scalar_one_or_none()

        if last_run and last_run.results:
            # имя листа — обрезаем до 31 символа (ограничение Excel)
            sheet_name = task.name[:31]
            # если лист с таким именем уже есть — добавляем суффикс
            existing_names = [s.title for s in wb.worksheets]
            if sheet_name in existing_names:
                sheet_name = sheet_name[:27] + f"_{len(existing_names)}"

            task_ws = wb.create_sheet(title=sheet_name)
            task_ws.append(["Поле", "Значение", "Индекс"])
            _style_header(task_ws)

            for r in last_run.results:
                task_ws.append([r.field_name, r.field_value or "", r.field_index])

            _auto_width(task_ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
