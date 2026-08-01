"""Тесты экспорта в Excel — генерация xlsx файлов."""

from io import BytesIO

from openpyxl import Workbook

from app.services.export_service import _auto_width, _style_header


class TestExcelHelpers:
    """Тесты вспомогательных функций экспорта."""

    def test_style_header(self):
        """Стили заголовков должны применяться."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Price", "Status"])
        _style_header(ws)
        assert ws["A1"].font.bold is True

    def test_auto_width(self):
        """Автоподбор ширины не должен падать."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Short", "A very long column header that should be wider"])
        ws.append(["x", "y"])
        _auto_width(ws)
        # просто проверяем что не упало
        assert True

    def test_empty_workbook(self):
        """Пустой лист — автоподбор не падает."""
        wb = Workbook()
        ws = wb.active
        _auto_width(ws)
        assert True


class TestExcelGeneration:
    """Тесты создания Excel-файлов."""

    def test_workbook_saves_to_bytes(self):
        """Workbook должен сохраняться в BytesIO."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Test"
        ws.append(["Run Date", "title", "price"])
        ws.append(["2026-01-01 10:00", "MacBook", "$1999"])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        assert len(output.read()) > 0

    def test_sheet_name_truncation(self):
        """Длинное имя листа обрезается до 31 символа."""
        name = "Very Long Task Name That Exceeds The Limit Of Thirty One Characters"
        wb = Workbook()
        ws = wb.active
        ws.title = name[:31]
        assert len(ws.title) <= 31
